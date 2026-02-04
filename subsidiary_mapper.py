"""
Corporate Subsidiary Mapper v2.0
Maps corporate structures using OpenCorporates and Companies House APIs.
Fixed Streamlit button handling with proper session state.
"""

import streamlit as st
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime
import time

# =============================================================================
# PAGE CONFIG - Must be first Streamlit command
# =============================================================================
st.set_page_config(
    page_title="Corporate Subsidiary Mapper",
    page_icon="🏢",
    layout="wide"
)

# =============================================================================
# SESSION STATE INITIALIZATION - Critical for button handling
# =============================================================================
if 'search_results' not in st.session_state:
    st.session_state.search_results = None
if 'selected_company' not in st.session_state:
    st.session_state.selected_company = None
if 'mapping_results' not in st.session_state:
    st.session_state.mapping_results = None
if 'search_query' not in st.session_state:
    st.session_state.search_query = ""

# =============================================================================
# API CONFIGURATION
# =============================================================================
OPENCORPORATES_BASE = "https://api.opencorporates.com/v0.4"
COMPANIES_HOUSE_BASE = "https://api.company-information.service.gov.uk"

# =============================================================================
# API FUNCTIONS
# =============================================================================

def search_opencorporates(company_name, jurisdiction=None):
    """Search OpenCorporates for companies matching the name."""
    params = {
        'q': company_name,
        'per_page': 20,
        'order': 'score'
    }
    if jurisdiction:
        params['jurisdiction_code'] = jurisdiction
    
    try:
        response = requests.get(
            f"{OPENCORPORATES_BASE}/companies/search",
            params=params,
            timeout=30
        )
        if response.status_code == 200:
            data = response.json()
            companies = data.get('results', {}).get('companies', [])
            return [c.get('company', {}) for c in companies]
        elif response.status_code == 429:
            st.warning("OpenCorporates rate limit reached. Please wait a moment.")
            return []
        else:
            return []
    except requests.exceptions.RequestException as e:
        st.error(f"API request failed: {str(e)}")
        return []

def get_company_details(jurisdiction_code, company_number):
    """Get detailed company information from OpenCorporates."""
    try:
        url = f"{OPENCORPORATES_BASE}/companies/{jurisdiction_code}/{company_number}"
        response = requests.get(url, timeout=30)
        if response.status_code == 200:
            return response.json().get('results', {}).get('company', {})
        return None
    except requests.exceptions.RequestException:
        return None

def search_subsidiaries_opencorporates(company_name, parent_jurisdiction=None):
    """
    Search for potential subsidiaries by looking for companies with similar names
    or containing the parent name.
    """
    subsidiaries = []
    search_terms = [
        company_name,
        f"{company_name} Ltd",
        f"{company_name} Limited",
        f"{company_name} Inc",
        f"{company_name} GmbH",
        f"{company_name} BV",
        f"{company_name} SA",
        f"{company_name} SAS"
    ]
    
    seen_companies = set()
    
    for term in search_terms[:3]:  # Limit to avoid rate limiting
        results = search_opencorporates(term)
        for company in results:
            company_id = f"{company.get('jurisdiction_code', '')}/{company.get('company_number', '')}"
            if company_id not in seen_companies:
                seen_companies.add(company_id)
                subsidiaries.append(company)
        time.sleep(0.5)  # Rate limiting protection
    
    return subsidiaries

def search_companies_house(company_name, api_key=None):
    """Search UK Companies House for companies."""
    headers = {}
    if api_key:
        headers['Authorization'] = f'Basic {api_key}'
    
    try:
        response = requests.get(
            f"{COMPANIES_HOUSE_BASE}/search/companies",
            params={'q': company_name, 'items_per_page': 20},
            headers=headers,
            timeout=30
        )
        if response.status_code == 200:
            return response.json().get('items', [])
        return []
    except requests.exceptions.RequestException:
        return []

def get_companies_house_officers(company_number, api_key=None):
    """Get officers (directors) for a UK company."""
    headers = {}
    if api_key:
        headers['Authorization'] = f'Basic {api_key}'
    
    try:
        response = requests.get(
            f"{COMPANIES_HOUSE_BASE}/company/{company_number}/officers",
            headers=headers,
            timeout=30
        )
        if response.status_code == 200:
            return response.json().get('items', [])
        return []
    except requests.exceptions.RequestException:
        return []

def get_companies_house_pscs(company_number, api_key=None):
    """Get Persons with Significant Control for a UK company."""
    headers = {}
    if api_key:
        headers['Authorization'] = f'Basic {api_key}'
    
    try:
        response = requests.get(
            f"{COMPANIES_HOUSE_BASE}/company/{company_number}/persons-with-significant-control",
            headers=headers,
            timeout=30
        )
        if response.status_code == 200:
            return response.json().get('items', [])
        return []
    except requests.exceptions.RequestException:
        return []

# =============================================================================
# DATA PROCESSING FUNCTIONS
# =============================================================================

def categorize_entities(entities, parent_name):
    """Categorize entities into parent, subsidiaries, and related companies."""
    parent_name_lower = parent_name.lower()
    
    categories = {
        'parent': [],
        'likely_subsidiaries': [],
        'related': [],
        'other': []
    }
    
    for entity in entities:
        name = entity.get('name', '').lower()
        
        # Check if this is the parent company
        if name == parent_name_lower or (
            parent_name_lower in name and 
            entity.get('company_type', '').lower() in ['plc', 'public limited company', 'holding company']
        ):
            categories['parent'].append(entity)
        # Check for clear subsidiary patterns
        elif parent_name_lower in name:
            categories['likely_subsidiaries'].append(entity)
        # Check for related company patterns
        elif any(word in name for word in parent_name_lower.split()[:2] if len(word) > 3):
            categories['related'].append(entity)
        else:
            categories['other'].append(entity)
    
    return categories

def create_excel_report(company_name, entities, categories):
    """Create a comprehensive Excel report of the corporate structure."""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    subheader_fill = PatternFill("solid", fgColor="BDD7EE")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- Summary Sheet ---
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = f"Corporate Structure Report: {company_name}"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:E1')
    
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws_summary['A2'].font = Font(italic=True, size=10)
    
    # Statistics
    ws_summary['A4'] = "Category"
    ws_summary['B4'] = "Count"
    ws_summary['A4'].font = header_font
    ws_summary['A4'].fill = header_fill
    ws_summary['B4'].font = header_font
    ws_summary['B4'].fill = header_fill
    
    row = 5
    for category, items in categories.items():
        ws_summary[f'A{row}'] = category.replace('_', ' ').title()
        ws_summary[f'B{row}'] = len(items)
        row += 1
    
    ws_summary[f'A{row}'] = "Total Entities"
    ws_summary[f'B{row}'] = len(entities)
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary[f'B{row}'].font = Font(bold=True)
    
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15
    
    # --- All Entities Sheet ---
    ws_all = wb.create_sheet("All Entities")
    
    headers = ["Company Name", "Jurisdiction", "Company Number", "Status", 
               "Company Type", "Incorporation Date", "Category", "OpenCorporates URL"]
    
    for col, header in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for category, items in categories.items():
        for entity in items:
            ws_all.cell(row=row, column=1, value=entity.get('name', 'N/A'))
            ws_all.cell(row=row, column=2, value=entity.get('jurisdiction_code', 'N/A'))
            ws_all.cell(row=row, column=3, value=entity.get('company_number', 'N/A'))
            ws_all.cell(row=row, column=4, value=entity.get('current_status', entity.get('company_status', 'N/A')))
            ws_all.cell(row=row, column=5, value=entity.get('company_type', 'N/A'))
            ws_all.cell(row=row, column=6, value=entity.get('incorporation_date', 'N/A'))
            ws_all.cell(row=row, column=7, value=category.replace('_', ' ').title())
            
            # OpenCorporates URL
            jur = entity.get('jurisdiction_code', '')
            num = entity.get('company_number', '')
            if jur and num:
                url = f"https://opencorporates.com/companies/{jur}/{num}"
                ws_all.cell(row=row, column=8, value=url)
            
            for col in range(1, 9):
                ws_all.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    # Adjust column widths
    widths = [50, 15, 20, 15, 25, 15, 20, 60]
    for i, width in enumerate(widths, 1):
        ws_all.column_dimensions[chr(64 + i)].width = width
    
    # --- Subsidiaries Sheet ---
    ws_subs = wb.create_sheet("Likely Subsidiaries")
    
    headers_subs = ["Company Name", "Jurisdiction", "Company Number", "Status", "Company Type"]
    
    for col, header in enumerate(headers_subs, 1):
        cell = ws_subs.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for entity in categories.get('likely_subsidiaries', []):
        ws_subs.cell(row=row, column=1, value=entity.get('name', 'N/A'))
        ws_subs.cell(row=row, column=2, value=entity.get('jurisdiction_code', 'N/A'))
        ws_subs.cell(row=row, column=3, value=entity.get('company_number', 'N/A'))
        ws_subs.cell(row=row, column=4, value=entity.get('current_status', entity.get('company_status', 'N/A')))
        ws_subs.cell(row=row, column=5, value=entity.get('company_type', 'N/A'))
        
        for col in range(1, 6):
            ws_subs.cell(row=row, column=col).border = thin_border
        row += 1
    
    widths_subs = [50, 15, 20, 15, 25]
    for i, width in enumerate(widths_subs, 1):
        ws_subs.column_dimensions[chr(64 + i)].width = width
    
    # --- By Jurisdiction Sheet ---
    ws_jur = wb.create_sheet("By Jurisdiction")
    
    # Group by jurisdiction
    by_jurisdiction = {}
    for entity in entities:
        jur = entity.get('jurisdiction_code', 'Unknown')
        if jur not in by_jurisdiction:
            by_jurisdiction[jur] = []
        by_jurisdiction[jur].append(entity)
    
    headers_jur = ["Jurisdiction", "Company Count", "Companies"]
    
    for col, header in enumerate(headers_jur, 1):
        cell = ws_jur.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row = 2
    for jur, companies in sorted(by_jurisdiction.items(), key=lambda x: -len(x[1])):
        ws_jur.cell(row=row, column=1, value=jur.upper())
        ws_jur.cell(row=row, column=2, value=len(companies))
        ws_jur.cell(row=row, column=3, value=", ".join([c.get('name', '')[:30] for c in companies[:5]]))
        
        for col in range(1, 4):
            ws_jur.cell(row=row, column=col).border = thin_border
        row += 1
    
    ws_jur.column_dimensions['A'].width = 15
    ws_jur.column_dimensions['B'].width = 15
    ws_jur.column_dimensions['C'].width = 100
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

# =============================================================================
# CALLBACK FUNCTIONS - Handle button clicks
# =============================================================================

def do_search():
    """Execute company search and store results in session state."""
    query = st.session_state.search_input
    if query:
        st.session_state.search_query = query
        with st.spinner(f"Searching for '{query}'..."):
            results = search_opencorporates(query)
            st.session_state.search_results = results
            st.session_state.selected_company = None
            st.session_state.mapping_results = None

def do_mapping():
    """Execute subsidiary mapping for selected company."""
    if st.session_state.selected_company:
        company = st.session_state.selected_company
        company_name = company.get('name', '')
        
        with st.spinner(f"Mapping subsidiaries for '{company_name}'..."):
            # Search for related entities
            entities = search_subsidiaries_opencorporates(company_name)
            
            # Add the parent company if not already included
            parent_id = f"{company.get('jurisdiction_code', '')}/{company.get('company_number', '')}"
            entity_ids = [f"{e.get('jurisdiction_code', '')}/{e.get('company_number', '')}" for e in entities]
            if parent_id not in entity_ids:
                entities.insert(0, company)
            
            # Categorize
            categories = categorize_entities(entities, company_name)
            
            st.session_state.mapping_results = {
                'company_name': company_name,
                'entities': entities,
                'categories': categories
            }

# =============================================================================
# MAIN UI
# =============================================================================

st.title("🏢 Corporate Subsidiary Mapper")
st.markdown("Map corporate structures and subsidiaries using OpenCorporates data")

# Sidebar
with st.sidebar:
    st.header("ℹ️ About")
    st.markdown("""
    This tool helps you:
    - Search for companies globally
    - Find related subsidiaries
    - Export corporate structure to Excel
    
    **Data Sources:**
    - OpenCorporates (global)
    - UK Companies House (coming soon)
    """)
    
    st.markdown("---")
    st.markdown("**Note:** OpenCorporates free tier has rate limits. For heavy use, consider their API subscription.")

# Main content
st.markdown("---")

# Search section
st.subheader("1️⃣ Search for a Company")

col1, col2 = st.columns([3, 1])

with col1:
    st.text_input(
        "Enter company name:",
        key="search_input",
        placeholder="e.g., Google, Shell, Volkswagen"
    )

with col2:
    st.markdown("<br>", unsafe_allow_html=True)
    st.button("🔍 Search", on_click=do_search, use_container_width=True)

# Display search results
if st.session_state.search_results is not None:
    results = st.session_state.search_results
    
    if not results:
        st.warning("No companies found. Try a different search term.")
    else:
        st.success(f"Found {len(results)} companies")
        
        st.markdown("---")
        st.subheader("2️⃣ Select a Company")
        
        # Create options for selectbox
        options = []
        for company in results:
            name = company.get('name', 'Unknown')
            jur = company.get('jurisdiction_code', '??').upper()
            status = company.get('current_status', company.get('company_status', 'Unknown'))
            option = f"{name} ({jur}) - {status}"
            options.append(option)
        
        selected_idx = st.selectbox(
            "Select a company to map:",
            range(len(options)),
            format_func=lambda x: options[x],
            key="company_selector"
        )
        
        # Store selected company
        if selected_idx is not None:
            st.session_state.selected_company = results[selected_idx]
            
            # Show company details
            company = results[selected_idx]
            with st.expander("Company Details", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    st.write(f"**Name:** {company.get('name', 'N/A')}")
                    st.write(f"**Jurisdiction:** {company.get('jurisdiction_code', 'N/A').upper()}")
                    st.write(f"**Company Number:** {company.get('company_number', 'N/A')}")
                with col2:
                    st.write(f"**Status:** {company.get('current_status', company.get('company_status', 'N/A'))}")
                    st.write(f"**Type:** {company.get('company_type', 'N/A')}")
                    st.write(f"**Incorporated:** {company.get('incorporation_date', 'N/A')}")
        
        # Map button
        st.markdown("---")
        st.subheader("3️⃣ Map Subsidiaries")
        
        st.button(
            "🗺️ Map Subsidiaries",
            on_click=do_mapping,
            use_container_width=True,
            type="primary"
        )

# Display mapping results
if st.session_state.mapping_results is not None:
    results = st.session_state.mapping_results
    
    st.markdown("---")
    st.subheader("4️⃣ Results")
    
    st.success(f"Found {len(results['entities'])} related entities for {results['company_name']}")
    
    # Summary metrics
    categories = results['categories']
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Parent/HQ", len(categories.get('parent', [])))
    with col2:
        st.metric("Likely Subsidiaries", len(categories.get('likely_subsidiaries', [])))
    with col3:
        st.metric("Related", len(categories.get('related', [])))
    with col4:
        st.metric("Other Matches", len(categories.get('other', [])))
    
    # Show entities by category
    tabs = st.tabs(["Likely Subsidiaries", "Related Companies", "All Entities"])
    
    with tabs[0]:
        subs = categories.get('likely_subsidiaries', [])
        if subs:
            for entity in subs:
                with st.container():
                    st.write(f"**{entity.get('name', 'N/A')}**")
                    st.caption(f"{entity.get('jurisdiction_code', '').upper()} | {entity.get('company_number', '')} | {entity.get('current_status', entity.get('company_status', 'N/A'))}")
        else:
            st.info("No clear subsidiaries identified. Check 'Related Companies' or 'All Entities' tabs.")
    
    with tabs[1]:
        related = categories.get('related', [])
        if related:
            for entity in related:
                with st.container():
                    st.write(f"**{entity.get('name', 'N/A')}**")
                    st.caption(f"{entity.get('jurisdiction_code', '').upper()} | {entity.get('company_number', '')} | {entity.get('current_status', entity.get('company_status', 'N/A'))}")
        else:
            st.info("No related companies found.")
    
    with tabs[2]:
        all_entities = results['entities']
        df = pd.DataFrame([{
            'Name': e.get('name', 'N/A'),
            'Jurisdiction': e.get('jurisdiction_code', '').upper(),
            'Number': e.get('company_number', 'N/A'),
            'Status': e.get('current_status', e.get('company_status', 'N/A')),
            'Type': e.get('company_type', 'N/A')
        } for e in all_entities])
        
        st.dataframe(df, use_container_width=True)
    
    # Export section
    st.markdown("---")
    st.subheader("5️⃣ Export")
    
    excel_buffer = create_excel_report(
        results['company_name'],
        results['entities'],
        results['categories']
    )
    
    filename = f"{results['company_name'].lower().replace(' ', '_')}_corporate_structure.xlsx"
    
    st.download_button(
        label="📥 Download Excel Report",
        data=excel_buffer,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# Footer
st.markdown("---")
st.caption("Data source: OpenCorporates. Note: Subsidiary identification is based on name matching and may require manual verification.")
