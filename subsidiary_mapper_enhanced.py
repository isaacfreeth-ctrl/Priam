import streamlit as st
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime
import time

st.set_page_config(page_title="Corporate Subsidiary Mapper", page_icon="🏢", layout="wide")

st.title("🏢 Corporate Subsidiary Mapper")
st.markdown("Map corporate structures using Companies House (UK) and OpenCorporates (Global)")

# API configuration
COMPANIES_HOUSE_BASE = "https://api.company-information.service.gov.uk"
OPENCORPORATES_BASE = "https://api.opencorporates.com/v0.4"

# Check for API keys
try:
    CH_API_KEY = st.secrets.get("COMPANIES_HOUSE_API_KEY", "")
    OC_API_KEY = st.secrets.get("OPENCORPORATES_API_KEY", "")
except:
    CH_API_KEY = ""
    OC_API_KEY = ""

def search_companies_house(company_name):
    """Search Companies House for UK companies"""
    if not CH_API_KEY:
        return []
    
    try:
        url = f"{COMPANIES_HOUSE_BASE}/search/companies"
        params = {'q': company_name, 'items_per_page': 10}
        auth = (CH_API_KEY, '')
        
        response = requests.get(url, params=params, auth=auth, timeout=30)
        if response.status_code == 200:
            data = response.json()
            return data.get('items', [])
        return []
    except Exception as e:
        st.error(f"Companies House search error: {str(e)}")
        return []

def get_ch_company_profile(company_number):
    """Get company profile from Companies House"""
    if not CH_API_KEY:
        return None
    
    try:
        url = f"{COMPANIES_HOUSE_BASE}/company/{company_number}"
        auth = (CH_API_KEY, '')
        response = requests.get(url, auth=auth, timeout=30)
        if response.status_code == 200:
            return response.json()
        return None
    except Exception as e:
        return None

def get_ch_officers(company_number):
    """Get company officers from Companies House"""
    if not CH_API_KEY:
        return []
    
    try:
        url = f"{COMPANIES_HOUSE_BASE}/company/{company_number}/officers"
        auth = (CH_API_KEY, '')
        response = requests.get(url, auth=auth, timeout=30)
        if response.status_code == 200:
            data = response.json()
            return data.get('items', [])
        return []
    except Exception as e:
        return []

def get_ch_filing_history(company_number):
    """Get filing history to find PSC statements and group structures"""
    if not CH_API_KEY:
        return []
    
    try:
        url = f"{COMPANIES_HOUSE_BASE}/company/{company_number}/filing-history"
        auth = (CH_API_KEY, '')
        params = {'category': 'capital', 'items_per_page': 50}
        response = requests.get(url, params=params, auth=auth, timeout=30)
        if response.status_code == 200:
            data = response.json()
            return data.get('items', [])
        return []
    except Exception as e:
        return []

def search_opencorporates(company_name, jurisdiction=None):
    """Search OpenCorporates (requires API key)"""
    params = {'q': company_name, 'per_page': 10}
    if jurisdiction:
        params['jurisdiction_code'] = jurisdiction
    if OC_API_KEY:
        params['api_token'] = OC_API_KEY
    
    try:
        response = requests.get(f"{OPENCORPORATES_BASE}/companies/search", params=params, timeout=30)
        if response.status_code == 200:
            data = response.json()
            return data.get('results', {}).get('companies', [])
        return []
    except Exception as e:
        return []

def search_by_officer_name(officer_name):
    """Find companies where an officer serves (may indicate group structure)"""
    if not CH_API_KEY:
        return []
    
    try:
        url = f"{COMPANIES_HOUSE_BASE}/search/officers"
        params = {'q': officer_name, 'items_per_page': 20}
        auth = (CH_API_KEY, '')
        
        response = requests.get(url, params=params, auth=auth, timeout=30)
        if response.status_code == 200:
            data = response.json()
            return data.get('items', [])
        return []
    except Exception as e:
        return []

def find_related_companies_by_officers(company_number, company_name):
    """Find related companies by looking at shared officers"""
    officers = get_ch_officers(company_number)
    related = []
    
    if not officers:
        return related
    
    # Get key officers (directors, corporate officers)
    key_officers = [o for o in officers if o.get('officer_role') in ['director', 'corporate-director']]
    
    for officer in key_officers[:3]:  # Limit to top 3 to avoid rate limits
        officer_name = officer.get('name', '')
        if not officer_name:
            continue
        
        time.sleep(0.6)  # Rate limiting
        
        officer_companies = search_by_officer_name(officer_name)
        for item in officer_companies:
            if item.get('company_number') != company_number:
                related.append({
                    'Level': 1,
                    'Parent Company': company_name,
                    'Company Name': item.get('title', 'Unknown'),
                    'Jurisdiction': 'gb',
                    'Company Number': item.get('company_number', ''),
                    'Status': item.get('company_status', 'Unknown'),
                    'Company Type': item.get('company_type', 'Unknown'),
                    'Incorporation Date': '',
                    'Connection Type': f'Shared officer: {officer_name}',
                    'Companies House URL': f"https://find-and-update.company-information.service.gov.uk/company/{item.get('company_number', '')}"
                })
    
    return related

def create_excel_export(subsidiary_data, root_company_name):
    """Create formatted Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Corporate Structure"
    
    headers = ['Level', 'Parent Company', 'Company Name', 'Jurisdiction', 'Company Number', 
               'Status', 'Company Type', 'Connection Type', 'URL']
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    body_font = Font(name='Arial', size=10)
    for row_idx, record in enumerate(subsidiary_data, 2):
        indent = "  " * record.get('Level', 0)
        company_name_display = f"{indent}{record['Company Name']}"
        
        ws.cell(row=row_idx, column=1, value=record.get('Level', 0)).font = body_font
        ws.cell(row=row_idx, column=2, value=record.get('Parent Company', '')).font = body_font
        ws.cell(row=row_idx, column=3, value=company_name_display).font = body_font
        ws.cell(row=row_idx, column=4, value=record.get('Jurisdiction', '')).font = body_font
        ws.cell(row=row_idx, column=5, value=record.get('Company Number', '')).font = body_font
        ws.cell(row=row_idx, column=6, value=record.get('Status', '')).font = body_font
        ws.cell(row=row_idx, column=7, value=record.get('Company Type', '')).font = body_font
        ws.cell(row=row_idx, column=8, value=record.get('Connection Type', 'Direct subsidiary')).font = body_font
        
        url = record.get('Companies House URL') or record.get('OpenCorporates URL', '')
        if url:
            url_cell = ws.cell(row=row_idx, column=9, value=url)
            url_cell.hyperlink = url
            url_cell.font = Font(color='0563C1', underline='single', name='Arial', size=10)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 50
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws['A1'] = "Corporate Structure Analysis"
    summary_ws['A1'].font = Font(bold=True, size=14, name='Arial')
    summary_ws['A3'] = "Root Company:"
    summary_ws['B3'] = root_company_name
    summary_ws['A4'] = "Total Entities:"
    summary_ws['B4'] = len(subsidiary_data)
    summary_ws['A5'] = "Generated:"
    summary_ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    return wb

# Main UI
st.sidebar.header("API Configuration")

col1, col2 = st.sidebar.columns(2)
with col1:
    ch_status = "✓" if CH_API_KEY else "✗"
    st.metric("Companies House", ch_status)
with col2:
    oc_status = "✓" if OC_API_KEY else "✗"
    st.metric("OpenCorporates", oc_status)

if not CH_API_KEY:
    st.sidebar.warning("Add Companies House API key for UK companies")
    with st.sidebar.expander("Get free API key"):
        st.markdown("[Companies House Developer Hub](https://developer.company-information.service.gov.uk/get-started)")

st.sidebar.header("Search Configuration")

search_source = st.sidebar.radio(
    "Data Source",
    ["Companies House (UK)", "OpenCorporates (Global)"],
    help="Companies House: Free, no auth required for public data | OpenCorporates: Requires paid API"
)

company_query = st.sidebar.text_input("Company Name", placeholder="e.g., Google UK, Unilever")

if search_source == "OpenCorporates (Global)":
    jurisdiction = st.sidebar.text_input("Jurisdiction (optional)", placeholder="e.g., gb, us_de")
else:
    jurisdiction = "gb"

search_button = st.sidebar.button("🔍 Search", type="primary")

if search_button and company_query:
    with st.spinner("Searching..."):
        if search_source == "Companies House (UK)":
            companies = search_companies_house(company_query)
            
            if companies:
                st.success(f"Found {len(companies)} UK companies")
                
                for idx, company in enumerate(companies):
                    col1, col2 = st.columns([4, 1])
                    
                    with col1:
                        st.markdown(f"**{company.get('title', 'Unknown')}**")
                        st.caption(f"Number: {company.get('company_number')} | Status: {company.get('company_status')}")
                    
                    with col2:
                        if st.button("Map", key=f"map_{idx}"):
                            with st.spinner("Building corporate structure..."):
                                company_number = company.get('company_number')
                                company_name = company.get('title')
                                
                                # Get company profile
                                profile = get_ch_company_profile(company_number)
                                
                                structure = [{
                                    'Level': 0,
                                    'Parent Company': 'ROOT',
                                    'Company Name': company_name,
                                    'Jurisdiction': 'gb',
                                    'Company Number': company_number,
                                    'Status': company.get('company_status', 'Unknown'),
                                    'Company Type': company.get('company_type', 'Unknown'),
                                    'Connection Type': 'Root entity',
                                    'Companies House URL': f"https://find-and-update.company-information.service.gov.uk/company/{company_number}"
                                }]
                                
                                # Find related companies by officers
                                st.info("🔍 Analyzing shared officers to identify potential group structure...")
                                related = find_related_companies_by_officers(company_number, company_name)
                                
                                if related:
                                    structure.extend(related)
                                    
                                    # Remove duplicates
                                    seen = set()
                                    unique_structure = []
                                    for item in structure:
                                        key = item['Company Number']
                                        if key not in seen:
                                            seen.add(key)
                                            unique_structure.append(item)
                                    
                                    st.success(f"Found {len(unique_structure)} related entities")
                                    
                                    df = pd.DataFrame(unique_structure)
                                    st.dataframe(df, use_container_width=True)
                                    
                                    wb = create_excel_export(unique_structure, company_name)
                                    excel_buffer = io.BytesIO()
                                    wb.save(excel_buffer)
                                    excel_buffer.seek(0)
                                    
                                    st.download_button(
                                        label="📥 Download Excel Report",
                                        data=excel_buffer,
                                        file_name=f"corporate_structure_{company_number}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                    )
                                    
                                    st.info("💡 Structure identified through shared officer analysis. This indicates potential group relationships but should be verified through official filings.")
                                else:
                                    st.warning("No related entities found through officer analysis. The company may be standalone or use different officer structures.")
                    
                    st.divider()
            else:
                st.warning("No companies found")
        
        else:  # OpenCorporates
            if not OC_API_KEY:
                st.error("OpenCorporates search requires an API key. Get one at [opencorporates.com/api_accounts/new](https://opencorporates.com/api_accounts/new)")
            else:
                companies = search_opencorporates(company_query, jurisdiction if jurisdiction else None)
                st.info(f"Found {len(companies)} companies (OpenCorporates data)")

else:
    st.info("👈 Enter a company name to begin")
    
    with st.expander("ℹ️ About This Tool"):
        st.markdown("""
        **Companies House (UK)**
        - Free API access with key
        - Real-time UK company data
        - Identifies relationships through shared officers
        - Official government source
        
        **OpenCorporates (Global)**
        - Requires paid API subscription  
        - Multi-jurisdiction coverage
        - Direct subsidiary data (when available)
        
        **Data Sources:**
        - [Companies House API](https://developer.company-information.service.gov.uk/)
        - [OpenCorporates API](https://api.opencorporates.com/documentation/)
        """)

st.divider()
st.caption("Built for European Political Influence Tracking")
