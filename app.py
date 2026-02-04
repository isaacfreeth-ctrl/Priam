"""
Corporate Network Mapper
Maps corporate relationships through shared directors/officers using Companies House data.
"""

import streamlit as st
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime
import base64
import time
from collections import defaultdict

# =============================================================================
# PAGE CONFIG
# =============================================================================
st.set_page_config(
    page_title="Corporate Network Mapper",
    page_icon="🕸️",
    layout="wide"
)

# =============================================================================
# API CONFIGURATION
# =============================================================================

def get_ch_api_key():
    """Get Companies House API key from secrets."""
    try:
        return st.secrets["COMPANIES_HOUSE_API_KEY"]
    except:
        return None

def ch_headers():
    """Get headers for Companies House API."""
    api_key = get_ch_api_key()
    if api_key:
        credentials = base64.b64encode(f"{api_key}:".encode()).decode()
        return {"Authorization": f"Basic {credentials}"}
    return {}

# =============================================================================
# COMPANIES HOUSE API FUNCTIONS
# =============================================================================

def search_companies(query):
    """Search for companies."""
    api_key = get_ch_api_key()
    if not api_key:
        st.error("Companies House API key not configured.")
        return []
    
    try:
        response = requests.get(
            "https://api.company-information.service.gov.uk/search/companies",
            params={"q": query, "items_per_page": 30},
            headers=ch_headers(),
            timeout=30
        )
        
        if response.status_code == 200:
            return response.json().get("items", [])
        elif response.status_code == 401:
            st.error("Invalid API key.")
            return []
        elif response.status_code == 429:
            st.warning("Rate limit hit. Wait a moment.")
            return []
        return []
    except Exception as e:
        st.error(f"Request failed: {e}")
        return []


def get_company_officers(company_number):
    """Get all officers for a company."""
    try:
        response = requests.get(
            f"https://api.company-information.service.gov.uk/company/{company_number}/officers",
            headers=ch_headers(),
            timeout=30
        )
        if response.status_code == 200:
            return response.json().get("items", [])
        return []
    except:
        return []


def get_officer_appointments(officer_id):
    """Get all appointments for an officer."""
    try:
        response = requests.get(
            f"https://api.company-information.service.gov.uk/officers/{officer_id}/appointments",
            headers=ch_headers(),
            timeout=30
        )
        if response.status_code == 200:
            return response.json().get("items", [])
        return []
    except:
        return []


def extract_officer_id(officer_link):
    """Extract officer ID from the links field."""
    # Link format: /officers/ABC123xyz/appointments
    if officer_link:
        parts = officer_link.strip("/").split("/")
        if len(parts) >= 2 and parts[0] == "officers":
            return parts[1]
    return None


def map_corporate_network(company_number, company_name, progress_callback=None):
    """
    Map corporate network through shared officers.
    Returns officers and all their connected companies.
    """
    # Step 1: Get officers of the target company
    officers = get_company_officers(company_number)
    
    if not officers:
        return None, None
    
    # Step 2: For each officer, get their other appointments
    officer_data = []
    connected_companies = defaultdict(lambda: {"company": None, "shared_officers": []})
    
    total = len(officers)
    
    for i, officer in enumerate(officers):
        if progress_callback:
            progress_callback((i + 1) / total, f"Checking officer {i+1}/{total}: {officer.get('name', 'Unknown')[:30]}...")
        
        officer_name = officer.get("name", "Unknown")
        officer_role = officer.get("officer_role", "Unknown")
        appointed = officer.get("appointed_on", "N/A")
        resigned = officer.get("resigned_on", None)
        
        # Skip resigned officers (optional - could include them)
        # if resigned:
        #     continue
        
        # Get officer ID from links
        officer_link = officer.get("links", {}).get("officer", {}).get("appointments", "")
        officer_id = extract_officer_id(officer_link)
        
        officer_info = {
            "name": officer_name,
            "role": officer_role,
            "appointed": appointed,
            "resigned": resigned,
            "officer_id": officer_id,
            "other_companies": []
        }
        
        if officer_id:
            # Get all their appointments
            appointments = get_officer_appointments(officer_id)
            time.sleep(0.2)  # Rate limiting
            
            for appt in appointments:
                appt_company_number = appt.get("appointed_to", {}).get("company_number", "")
                appt_company_name = appt.get("appointed_to", {}).get("company_name", "Unknown")
                appt_role = appt.get("officer_role", "Unknown")
                appt_status = appt.get("appointed_to", {}).get("company_status", "unknown")
                
                # Skip the source company
                if appt_company_number == company_number:
                    continue
                
                officer_info["other_companies"].append({
                    "company_number": appt_company_number,
                    "company_name": appt_company_name,
                    "role": appt_role,
                    "status": appt_status
                })
                
                # Track connected companies
                if appt_company_number:
                    connected_companies[appt_company_number]["company"] = {
                        "number": appt_company_number,
                        "name": appt_company_name,
                        "status": appt_status
                    }
                    connected_companies[appt_company_number]["shared_officers"].append({
                        "name": officer_name,
                        "role_at_source": officer_role,
                        "role_at_connected": appt_role
                    })
        
        officer_data.append(officer_info)
    
    return officer_data, dict(connected_companies)


def create_excel(source_company, officers, connected_companies):
    """Create comprehensive Excel report."""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    highlight_fill = PatternFill("solid", fgColor="BDD7EE")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # --- Summary Sheet ---
    ws = wb.active
    ws.title = "Summary"
    
    ws['A1'] = f"Corporate Network Analysis"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Source Company: {source_company['name']}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A3'] = f"Company Number: {source_company['number']}"
    ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    ws['A6'] = "Key Metrics"
    ws['A6'].font = Font(bold=True)
    ws['A7'] = "Total Officers:"
    ws['B7'] = len(officers)
    ws['A8'] = "Connected Companies:"
    ws['B8'] = len(connected_companies)
    
    # Count by shared officer count
    by_shared = defaultdict(int)
    for comp_data in connected_companies.values():
        count = len(comp_data["shared_officers"])
        by_shared[count] += 1
    
    ws['A10'] = "Companies by Shared Officer Count"
    ws['A10'].font = Font(bold=True)
    row = 11
    for count in sorted(by_shared.keys(), reverse=True):
        ws[f'A{row}'] = f"{count} shared officer(s):"
        ws[f'B{row}'] = by_shared[count]
        row += 1
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    # --- Connected Companies Sheet ---
    ws2 = wb.create_sheet("Connected Companies")
    
    headers = ["Company Name", "Company Number", "Status", "Shared Officers", "Officer Names", "Companies House URL"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Sort by number of shared officers (descending)
    sorted_companies = sorted(
        connected_companies.items(),
        key=lambda x: len(x[1]["shared_officers"]),
        reverse=True
    )
    
    row = 2
    for comp_number, comp_data in sorted_companies:
        company = comp_data["company"]
        shared = comp_data["shared_officers"]
        
        ws2.cell(row=row, column=1, value=company["name"] if company else "Unknown")
        ws2.cell(row=row, column=2, value=comp_number)
        ws2.cell(row=row, column=3, value=company["status"] if company else "Unknown")
        ws2.cell(row=row, column=4, value=len(shared))
        ws2.cell(row=row, column=5, value=", ".join([s["name"] for s in shared]))
        ws2.cell(row=row, column=6, value=f"https://find-and-update.company-information.service.gov.uk/company/{comp_number}")
        
        # Highlight companies with multiple shared officers
        if len(shared) >= 2:
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = highlight_fill
        
        for col in range(1, 7):
            ws2.cell(row=row, column=col).border = border
        
        row += 1
    
    ws2.column_dimensions['A'].width = 50
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 50
    ws2.column_dimensions['F'].width = 60
    
    # --- Officers Sheet ---
    ws3 = wb.create_sheet("Officers")
    
    headers = ["Officer Name", "Role", "Appointed", "Resigned", "Other Directorships"]
    for col, h in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row = 2
    for officer in officers:
        ws3.cell(row=row, column=1, value=officer["name"])
        ws3.cell(row=row, column=2, value=officer["role"])
        ws3.cell(row=row, column=3, value=officer["appointed"])
        ws3.cell(row=row, column=4, value=officer["resigned"] or "Current")
        ws3.cell(row=row, column=5, value=len(officer["other_companies"]))
        
        for col in range(1, 6):
            ws3.cell(row=row, column=col).border = border
        
        row += 1
    
    ws3.column_dimensions['A'].width = 35
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 18
    
    # --- Detailed Connections Sheet ---
    ws4 = wb.create_sheet("Detailed Connections")
    
    headers = ["Officer Name", "Role at Source", "Connected Company", "Role at Connected", "Company Number", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    row = 2
    for officer in officers:
        for other in officer["other_companies"]:
            ws4.cell(row=row, column=1, value=officer["name"])
            ws4.cell(row=row, column=2, value=officer["role"])
            ws4.cell(row=row, column=3, value=other["company_name"])
            ws4.cell(row=row, column=4, value=other["role"])
            ws4.cell(row=row, column=5, value=other["company_number"])
            ws4.cell(row=row, column=6, value=other["status"])
            
            for col in range(1, 7):
                ws4.cell(row=row, column=col).border = border
            
            row += 1
    
    ws4.column_dimensions['A'].width = 35
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 50
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 15
    ws4.column_dimensions['F'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =============================================================================
# MAIN APP
# =============================================================================

st.title("🕸️ Corporate Network Mapper")
st.markdown("Map corporate relationships through shared directors and officers")

# Check for API key
if not get_ch_api_key():
    st.error("⚠️ Companies House API key not found!")
    st.markdown("""
    **To fix this:**
    1. Go to Streamlit Cloud app settings
    2. Click **Secrets**
    3. Add:
    ```
    COMPANIES_HOUSE_API_KEY = "your_api_key_here"
    ```
    """)
    st.stop()

# Sidebar
with st.sidebar:
    st.header("How it works")
    st.markdown("""
    1. **Search** for a company
    2. **Select** the target company
    3. **Map** finds all officers/directors
    4. For each officer, finds **all their other directorships**
    5. Shows **connected companies** ranked by shared officers
    
    **Why this works:**
    - Parent company directors often sit on subsidiary boards
    - Shared directors reveal corporate group structures
    - Catches connections that name matching misses
    """)
    
    st.markdown("---")
    st.markdown("✅ Companies House API connected")
    st.markdown("🇬🇧 UK companies only")

st.markdown("---")

# Step 1: Search
st.subheader("1. Search for a company")
search_term = st.text_input("Company name:", placeholder="e.g., Tesco, Barclays, BP")

if st.button("🔍 Search", type="primary"):
    if search_term:
        with st.spinner("Searching..."):
            results = search_companies(search_term)
        
        if results:
            st.session_state['search_results'] = results
            st.success(f"Found {len(results)} companies")
        else:
            st.warning("No companies found.")
            if 'search_results' in st.session_state:
                del st.session_state['search_results']
    else:
        st.warning("Enter a company name")

# Step 2: Select
if 'search_results' in st.session_state and st.session_state['search_results']:
    st.markdown("---")
    st.subheader("2. Select a company")
    
    results = st.session_state['search_results']
    
    options = []
    for c in results:
        name = c.get("title", "Unknown")
        number = c.get("company_number", "")
        status = c.get("company_status", "?")
        options.append(f"{name} ({number}) - {status}")
    
    selection = st.selectbox("Choose:", options)
    selected_idx = options.index(selection)
    selected = results[selected_idx]
    
    with st.expander("Company details", expanded=True):
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**Name:** {selected.get('title', 'N/A')}")
            st.write(f"**Number:** {selected.get('company_number', 'N/A')}")
        with col2:
            st.write(f"**Status:** {selected.get('company_status', 'N/A')}")
            st.write(f"**Type:** {selected.get('company_type', 'N/A')}")
    
    # Step 3: Map
    st.markdown("---")
    st.subheader("3. Map corporate network")
    
    st.info("⏱️ This may take 1-2 minutes depending on the number of officers.")
    
    if st.button("🕸️ Map Network", type="primary"):
        company_number = selected.get("company_number", "")
        company_name = selected.get("title", "")
        
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        def update_progress(pct, msg):
            progress_bar.progress(pct)
            status_text.text(msg)
        
        officers, connected = map_corporate_network(
            company_number, 
            company_name,
            progress_callback=update_progress
        )
        
        progress_bar.empty()
        status_text.empty()
        
        if officers and connected:
            st.session_state['mapping_results'] = {
                'source': {'name': company_name, 'number': company_number},
                'officers': officers,
                'connected': connected
            }
        else:
            st.warning("No officers found for this company.")

# Step 4: Results
if 'mapping_results' in st.session_state and st.session_state['mapping_results']:
    st.markdown("---")
    st.subheader("4. Results")
    
    data = st.session_state['mapping_results']
    source = data['source']
    officers = data['officers']
    connected = data['connected']
    
    st.success(f"**{source['name']}** has **{len(officers)} officers** connected to **{len(connected)} other companies**")
    
    # Metrics
    multi_shared = sum(1 for c in connected.values() if len(c["shared_officers"]) >= 2)
    
    col1, col2, col3 = st.columns(3)
    col1.metric("Officers", len(officers))
    col2.metric("Connected Companies", len(connected))
    col3.metric("Strong Links (2+ shared)", multi_shared)
    
    # Tabs
    tab1, tab2, tab3 = st.tabs(["Connected Companies", "Officers", "Network Stats"])
    
    with tab1:
        # Sort by shared officer count
        sorted_conn = sorted(
            connected.items(),
            key=lambda x: len(x[1]["shared_officers"]),
            reverse=True
        )
        
        st.markdown("**Companies sorted by number of shared officers:**")
        
        for comp_num, comp_data in sorted_conn[:30]:  # Show top 30
            company = comp_data["company"]
            shared = comp_data["shared_officers"]
            
            shared_count = len(shared)
            indicator = "🔴" if shared_count >= 3 else "🟡" if shared_count >= 2 else "⚪"
            
            company_name = company["name"] if company else "Unknown"
            status = company["status"] if company else "unknown"
            
            st.write(f"{indicator} **{company_name}** ({shared_count} shared)")
            st.caption(f"{comp_num} | {status} | Shared: {', '.join([s['name'] for s in shared][:3])}{'...' if len(shared) > 3 else ''}")
        
        if len(sorted_conn) > 30:
            st.info(f"Showing top 30 of {len(sorted_conn)} connected companies. Download Excel for full list.")
    
    with tab2:
        for officer in officers:
            other_count = len(officer["other_companies"])
            status = "Current" if not officer["resigned"] else f"Resigned {officer['resigned']}"
            
            st.write(f"**{officer['name']}** - {officer['role']}")
            st.caption(f"{status} | {other_count} other directorships")
    
    with tab3:
        st.markdown("**Distribution of shared officers:**")
        
        by_count = defaultdict(list)
        for comp_num, comp_data in connected.items():
            count = len(comp_data["shared_officers"])
            by_count[count].append(comp_data["company"]["name"] if comp_data["company"] else comp_num)
        
        for count in sorted(by_count.keys(), reverse=True):
            companies = by_count[count]
            st.write(f"**{count} shared officer(s):** {len(companies)} companies")
            if count >= 2:
                for name in companies[:5]:
                    st.caption(f"  • {name}")
                if len(companies) > 5:
                    st.caption(f"  • ... and {len(companies) - 5} more")
    
    # Export
    st.markdown("---")
    st.subheader("5. Export")
    
    excel_data = create_excel(source, officers, connected)
    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in source['name'])
    filename = f"{safe_name}_corporate_network.xlsx"
    
    st.download_button(
        "📥 Download Excel Report",
        data=excel_data,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    st.markdown("""
    **Excel contains:**
    - Summary with key metrics
    - Connected Companies (sorted by shared officers, highlighted for 2+)
    - Officers list with directorship counts
    - Detailed connections (every officer ↔ company link)
    """)

# Footer
st.markdown("---")
st.caption("Data: UK Companies House | Network mapping via shared officers/directors")
